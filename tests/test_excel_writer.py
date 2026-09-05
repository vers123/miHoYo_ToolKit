"""extractors.excel_writer 单元测试

用临时 db 与临时 .xlsx，验证三 sheet 生成、数据行、空库、覆盖、冻结首行。
"""

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from core.storage import NewsStorage
from extractors.excel_writer import export_to_excel


class TestExcelWriter(unittest.TestCase):
    def setUp(self):
        self._db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        self._db.close()
        self._out = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        self._out.close()
        Path(self._out.name).unlink()  # 让 export 新建文件
        self.store = NewsStorage(db_path=self._db.name)

    def tearDown(self):
        self.store.close()
        Path(self._db.name).unlink(missing_ok=True)
        Path(self._out.name).unlink(missing_ok=True)

    def _item(self, info_id):
        return {
            "iInfoId": str(info_id),
            "sTitle": f"标题{info_id}",
            "date": "2024-01-01",
            "sCategoryName": "公告",
            "sIntro": "摘要",
            "poster_url": "http://img/x.jpg",
            "url": f"http://example.com/{info_id}",
        }

    def test_export_creates_three_sheets(self):
        self.store.upsert_items("genshin", [self._item(1)])
        self.store.upsert_items("zzz", [self._item(2)])
        self.store.upsert_items("starrail", [self._item(3)])
        out = export_to_excel(output_path=self._out.name, db_path=self._db.name)
        self.assertTrue(out.exists())
        wb = load_workbook(out)
        self.assertIn("原神", wb.sheetnames)
        self.assertIn("绝区零", wb.sheetnames)
        self.assertIn("星穹铁道", wb.sheetnames)

    def test_export_data_rows_and_header(self):
        self.store.upsert_items("genshin", [self._item(1), self._item(2)])
        export_to_excel(output_path=self._out.name, db_path=self._db.name)
        wb = load_workbook(self._out.name)
        ws = wb["原神"]
        # 表头 + 2 数据行
        self.assertEqual(ws.max_row, 3)
        self.assertEqual(ws.cell(1, 1).value, "ID")  # 表头
        self.assertEqual(ws.cell(2, 1).value, "1")  # 数据 iInfoId
        self.assertEqual(ws.cell(2, 2).value, "标题1")  # sTitle

    def test_export_empty_db(self):
        """空库也生成 4 个空 sheet（仅表头）"""
        out = export_to_excel(output_path=self._out.name, db_path=self._db.name)
        wb = load_workbook(out)
        self.assertEqual(len(wb.sheetnames), 4)
        for name in ("原神", "原神(EN)", "绝区零", "星穹铁道"):
            self.assertEqual(wb[name].max_row, 1)  # 仅表头

    def test_export_overwrites_existing_sheet(self):
        """二次导出覆盖旧 sheet，不产生重复 sheet"""
        self.store.upsert_items("genshin", [self._item(1)])
        export_to_excel(output_path=self._out.name, db_path=self._db.name)
        self.store.upsert_items("genshin", [self._item(2)])
        export_to_excel(output_path=self._out.name, db_path=self._db.name)
        wb = load_workbook(self._out.name)
        ws = wb["原神"]
        self.assertEqual(ws.max_row, 3)  # 表头 + 2 条
        self.assertEqual(wb.sheetnames.count("原神"), 1)  # 无重复

    def test_freeze_panes(self):
        self.store.upsert_items("genshin", [self._item(1)])
        export_to_excel(output_path=self._out.name, db_path=self._db.name)
        wb = load_workbook(self._out.name)
        self.assertEqual(wb["原神"].freeze_panes, "A2")

    def test_selective_games(self):
        """只导出部分游戏"""
        self.store.upsert_items("genshin", [self._item(1)])
        export_to_excel(
            games=["genshin"], output_path=self._out.name, db_path=self._db.name
        )
        wb = load_workbook(self._out.name)
        self.assertEqual(wb.sheetnames, ["原神"])


if __name__ == "__main__":
    unittest.main()
