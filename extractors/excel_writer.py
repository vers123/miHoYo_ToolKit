"""新闻 Excel 导出（D4）

从 SQLite 读取三游戏新闻，写入 .xlsx，每个游戏一个 sheet。
表头蓝底白字，冻结首行，按日期倒序。已存在文件则覆盖对应游戏 sheet
（保留其他游戏数据，便于增量再导出）。
"""

from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.storage import GAME_TABLES, NewsStorage
from utils.logger import get_module_logger

logger = get_module_logger(__name__)

_DEFAULT_OUTPUT = Path("output") / "news_export.xlsx"

# 列定义：(表头名, 字段名, 列宽)
_COLUMNS = [
    ("ID", "iInfoId", 10),
    ("标题", "sTitle", 40),
    ("日期", "date", 20),
    ("分类", "sCategoryName", 12),
    ("摘要", "sIntro", 50),
    ("封面", "poster_url", 30),
    ("链接", "url", 50),
]

_SHEET_NAMES = {"genshin": "原神", "zzz": "绝区零", "starrail": "星穹铁道"}

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def export_to_excel(
    games: Optional[List[str]] = None,
    output_path: Optional[str] = None,
    db_path: Optional[str] = None,
) -> Path:
    """从 SQLite 导出三游戏新闻到 Excel（每游戏一 sheet）

    Args:
        games: 要导出的游戏列表，默认全部三游戏
        output_path: 输出 .xlsx 路径，默认 output/news_export.xlsx
        db_path: SQLite 路径，默认 data/news.db

    Returns:
        输出文件路径
    """
    if games is None:
        games = list(GAME_TABLES.keys())
    out = Path(output_path) if output_path else _DEFAULT_OUTPUT
    out.parent.mkdir(parents=True, exist_ok=True)

    with NewsStorage(db_path=db_path) as store:
        # 追加模式：已存在则加载（保留其他 sheet），否则新建
        wb = load_workbook(out) if out.exists() else Workbook()
        if not out.exists() and "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        total = 0
        for game in games:
            sheet_name = _SHEET_NAMES.get(game, game)
            # 覆盖该游戏的旧 sheet
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            total += _write_sheet(ws, store, game)

        wb.save(out)

    logger.info(f"Excel 已导出: {out}（{len(games)} 个 sheet，共 {total} 条）")
    print(f"\n[OK] Excel 导出完成: {out}")
    print(f"     共 {len(games)} 个 sheet，{total} 条新闻")
    return out


def _write_sheet(ws, store: NewsStorage, game: str) -> int:
    """写单个游戏 sheet，返回写入条数"""
    # 表头
    for col, (title, _, _) in enumerate(_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # 数据（get_all_items 已按日期倒序）
    items = store.get_all_items(game)
    for r_idx, item in enumerate(items, 2):
        for c_idx, (_, field, _) in enumerate(_COLUMNS, 1):
            ws.cell(row=r_idx, column=c_idx, value=item.get(field, ""))

    # 列宽
    for i, (_, _, width) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 冻结首行，方便滚动查看
    ws.freeze_panes = "A2"

    return len(items)


def run(games: Optional[List[str]] = None, output_path: Optional[str] = None) -> Path:
    """命令行入口（与 extractors 其他 run_xxx 风格一致）"""
    return export_to_excel(games=games, output_path=output_path)
